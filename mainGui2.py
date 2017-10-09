'''
File Name:			mainGui2.py

Engineer(s):		Dustin Henderson

Last Edit:			10/9/17

Short Description:	This script is used to analise the analytics files from the cloud.altera.com design store. The script presents 
*					the options to the user in an easy to use gui. There are two main functions in the script. The first function
*					reads the input file and creates new excell sheets organising the data. This includes the number of downloads
*					by catigory, design, device family, and development kit. The second function reads the input file for 
*					documentation links. The links are then checked. If the link is dead or no documentation exsists for the 
*					project the referance design name and are added to a new sheet of the output xlsx file.
'''

print "***** Starting Up *****"
try:
	import time
except ImportError :
	print 'ERROR : Run the batch file again \nOR \nMake sure \'time\' package is installed correctly, to check enter in command prompt: \n<python -m pip install time> and re-run the script'
	exit()	
try:
	import threading
except ImportError :
	print 'ERROR : Run the batch file again \nOR \nMake sure \'threading\' package is installed correctly, to check enter in command prompt: \n<python -m pip install threading> and re-run the script'
	exit()	

'''***********************************************************************************************************************'''
'''****************************************** IMPORTS FOR ANALYTICS ******************************************************'''
'''***********************************************************************************************************************'''
try :
	import openpyxl
	from openpyxl.styles import Font
	from openpyxl import load_workbook
except ImportError :
	print 'ERROR : Run the batch file again \nOR \nMake sure \'openpyxl\' package is installed correctly, to check enter in command prompt: \n<python -m pip install openpyxl> and re-run the script'
	exit()	
try :
	import glob
except ImportError :
	print 'ERROR : Run the batch file again \nOR \nMake sure \'glob\' package is installed correctly, to check enter in command prompt: \n<python -m pip install glob>'
	exit()	
try :
	import os
except ImportError :
	print 'ERROR : Run the batch file again \nOR \nMake sure \'os\' package is installed correctly, to check enter in command prompt: \n<python -m pip install os>'
	exit()	
try :
	import requests
except ImportError :
	print 'ERROR : Run the batch file again \nOR \nMake sure \'requests\' package is installed correctly, to check enter in command prompt: \n<python -m pip install requests>'
	exit()

'''***********************************************************************************************************************'''
'''********************************************* IMPORTS FOR GUI *********************************************************'''
'''***********************************************************************************************************************'''
try:
	from Tkinter import *
	import Tkinter, Tkconstants, tkFileDialog
except:
	print 'ERROR : Run the batch file again \nOR \nMake sure \'Tkinter\' package is installed correctly, to check enter in command prompt: \n<python -m pip install python-tk>'
	exit()

'''***********************************************************************************************************************'''
'''**************************************** IMPORTS FOR CHEKING LINKS ****************************************************'''
'''***********************************************************************************************************************'''	
try:
	from pypac import PACSession, get_pac
except ImportError :
	print 'ERROR : Run the batch file again \nOR \nMake sure \'pypac\' package is installed correctly, to check enter in command prompt: \n<python -m pip install pypac>'
	exit()
try:
	import string
except ImportError :
	print 'ERROR : Run the batch file again \nOR \nMake sure \'ImportError\' package is installed correctly, to check enter in command prompt: \n<python -m pip install ImportError>'
	exit()
	
'''***********************************************************************************************************************'''
'''**************************************** IMPORTS FOR INTERNAL DEFS ****************************************************'''
'''***********************************************************************************************************************'''
try:
	import csv
except:
	print 'ERROR : Run the batch file again \nOR \nMake sure \'csv\' package is installed correctly, to check enter in command prompt: \n<python -m pip install csv>'
	exit()

try:
	import datetime
except:
	print 'ERROR : Run the batch file again \nOR \nMake sure \'datetime\' package is installed correctly, to check enter in command prompt: \n<python -m pip install datetime>'
	exit()

'''***********************************************************************************************************************'''
'''************************************ DEF DECLORATIONS AND IMPORTS *****************************************************'''
'''***********************************************************************************************************************'''

from analyticsCopy2 import runAnalytics


'''
Def Name:		convertToXlsx
*
Arguments:		inputLoc 	file path and name of the input csv file.
*				outputLoc	file path and name of the output xlsx file (if the xlsx file does not exsist it will be created)
*
Returns:		returns a bool to let the foallowing functions know if the convertion was succeffuly completed.
*
Description:	This def opens the csv file and reads it line by line into a xlsx file then saves the file xlsx file.
*				After the csv is read the csv is closed and the xlsx file is saved to the hard drive
'''
def convertToXlsx(inputLoc, outputLoc):
	try:
		wb = openpyxl.Workbook()						#create a workBook
		ws = wb.active									#set it as the active file
		csvFile = open(inputLoc)						#open the original CSV file with the orignal data from the site
		reader = csv.reader(csvFile, delimiter=',')		#read the data in the csv. currently the delimiter used is a ,
		for row in reader:								#read untill the bottom row
			ws.append(row)								#append each row to the active xlsx file
		csvFile.close()									#close the CSV file to prevent corruption
		wb.save(outputLoc)								#save the xlsx file in the output location
		return True										#Return true to notify that the process was completed
	except:
		return False									#Return False if the convertion failed

'''
Def Name:		copyXLSX
*
Arguments:		inputLoc 	file path and name of the input xlsx file.
*				outputLoc	file path and name of the output xlsx file (if the xlsx file does not exsist it will be created)
*
Returns:		returns a bool to let the foallowing functions know if the convertion was succeffuly completed.
*
Description:	This def opens the input xlsx file renames and copies it to the output directory. 
'''
def copyXLSX(inputLoc, outputLoc):						
	try:
		wbRead = load_workbook(filename = inputLoc)		#read then input xlsx
		wsRead = wbRead.active							#set the read xlsx file as active
		wbRead.save(outputLoc)							#save the file with the output name and location
		return True										#return true when this is succeffuly completed
	except:
		return False									#else return false if it fails

#Probs going to delete this soon...		
def testFunctionTime(bool):								#test function
	timeCount = 0										
	print "Thread start: ", bool.isSet()				
	while(timeCount < 10):								
		time.sleep(1)									
		print "check", timeCount, "threads: ", bool.isSet()
		timeCount = timeCount + 1						
		if(bool.isSet()):								
			break										

'''***********************************************************************************************************************'''
'''******************************************** LINK CHECK LOGIC *********************************************************'''
'''***********************************************************************************************************************'''
			
def linkCheck(fileLoc, bool):	
	class linkStuff:
		def __init__ (com):
			print"initial"	#used for testing will probably remove soon
			com.startThredCount = threading.activeCount()											#figure the startin 
			com.documentation = []																	#initial com.documentation. This is used to store all the 
			"""Removed the proxy line!!! get this from dustin"""																						#set up the pac session by getting the proxy wpad file
			com.session = PACSession(com.pac)														#set the session with the proxy
			com.testCounter = 0																	#delete this later
			com.resultList = []																		#initial com.resultList. This is used to store all the design examples and links that fail testing.
			com.getLinks(fileLoc)																	#run the getLinks def to get all the documentation links from the xlsx file
			for row in com.documentation:															#for each link in the documentation list check the link
				for x in range(1, 6): 																#check each link individually. there are 5 possible links for each design example
					if(row[x][0] != "-"):															#check if the link is empty
						tman = threading.Thread(target = com.requesLink, args = (row[0], row[x]))	#spin off a new thread for each link. This keeps the slow to respond servers from holduing up the fast ones
						tman.start()																#start the new thread created for the link
				com.testCounter = com.testCounter + 1
				'''**********************************************************************************************'''
				if(bool.isSet()):	#if the abort button is hit the bool breaks the loop for the threads								
					break # <====================== break in the loop
				'''**********************************************************************************************'''
				print threading.activeCount()														
				if(threading.activeCount() > (100 + com.startThredCount)):							#if too many threads get started wait untill some of them complete and close out
					while(threading.activeCount() > (com.startThredCount + 20)):					#while there are more than 20 threads loop though the sleep loop and recheck how many threads are running
						time.sleep(1)																#sleep for one second. all the other threads are running their connections. This keeps the system from spinning off new threads untill some finish and close
			while(threading.activeCount() > com.startThredCount):									#before the results list are appended to the xlsx file wait for all the threads checking links to finish and close
				time.sleep(1)
				print threading.activeCount()
				'''**********************************************************************************************'''
				if(bool.isSet()):	#if the abort button is hit the bool breaks the loop for the threads
					break # <====================== break in the loop
				'''**********************************************************************************************'''
			com.writeResults()	#use the writeResults def to save and dead links or examples that have problems to the xlsx file
		
		'''
		Def Name:	removeNonASCII
		*
		Arguments:		com 	com contains the varibles with in the class linkStuff
		*				string	string contains the string that will be striped of any non ASCII charicters
		*
		Returns:		returns the input string striped of any ASCII varibles
		*
		Description:	This is an extreemly simple def that runs through a string charicter by charicter and removes
		*				all ASCII charicters fall with in 0-127. The string is then put back together with the join
		*				and returned.
		'''
		def removeNonASCII(com, string):
			stripped = (c for c in string if 0 < ord(c) < 127)	#if the charicter falls outside of 0-127 delete it
			return ''.join(stripped)							#return the string re joined with any non ascii letter removed
		
		'''
		Def Name:	condenceUrls
		*
		Arguments:		com 	com contains the varibles with in the class linkStuff
		*
		Returns:		NA
		*
		Description:	This def simply looks for duplicate list items. Because the list it is working on is multidimentional
		*				it convirts
		'''
		def condenceUrls(com):
			com.documentation = set(tuple(element) for element in com.documentation)						#remove any duplicates in the list
			com.documentation = [list(t) for t in set(tuple(element) for element in com.documentation)]		#convert back to list for the return
			com.resultList = set(tuple(element) for element in com.resultList)								#remove any duplicates in the list
			com.resultList = [list(t) for t in set(tuple(element) for element in com.resultList)]			#convert back to list for the return
		
		'''
		Def Name:	findColumnLetter
		*
		Arguments:		com 	com contains the varibles with in the class linkStuff
		*				cell	the cell is the input value that the dictionary look up find the matching title in				
		*
		Returns:		str(cell)[start:end]	this is a single letter that corisponds to the columb that the title is located in
		*
		Description:	This def loops though the sting returned by the dictionary match. The string looks like <Cell .A1>. The loop
		*				loops though this and finds the letter that the corasponds to the column the match was found in.
		'''
		def findColumnLetter(com, cell):
			print "Charicter List"					#was a line used for testing. not important
			start = 0								#initial the start at 0. This is used to store starting location of the colum letter(s)
			end = 0									#initial the end at 0. This is used to store the ending location of the colum letter(s)
			charCounter = 0							#counter used to inrement while the for loop exicutes.
			for charicter in str(cell):				#loop through all the charicter in the string
				if(charicter == "."):				#the . indicate that the colum letter is the next charicter.
					start = charCounter + 1			#save the location for the start
					print "start ", start			
				if(charicter == "1"):				#the title is always going to be in the first row. the 1 indicates the end of the colum letter.
					end = charCounter				#save the lcation for the end of the colum charicter
					print "end", end				
				charCounter = charCounter + 1		#increment the counter as the for loop goes through each charicter
			print str(cell)[start:end]				
			return str(cell)[start:end]				#return only the letter for the column.
		
		'''
		Def Name:	getLinks
		*
		Arguments:		com 	com contains the varibles with in the class linkStuff
		*				fileLoc	this is the file path and name of the xlsx file convirted from the original csv file			
		*
		Returns:		NA 		all changes are done under the class
		*
		Description:	This def starts by opening the file location passed to it with openpyxl. After the file is opend the
		*				def starts to look for the columns containing the name of the design examples and the documentation
		*				when it looks for a match it runs the string though a filter removing whitespace and changing uppercase
		*				letters to lowercase. After the coumns are found the columns are read and appended to a 
		*				multidimentional list com.documentation. This while the def reads the xlsx file it also looks for 
		*				design examples that have not documentation. The designs that do not have any documentation are
		*				appended to the com.resultList. The com.resultList is used to save referance designs that have
		*				documentation problems.
		'''
		def getLinks(com, fileLoc):
			print fileLoc
			nameCol = ""		#used to store what columb the name of the example designs are stored in
			doc1Col = ""		#used to store what columb the first doc link is in
			doc2Col = ""		#used to store what columb the second doc link is in
			doc3Col = ""		#used to store what columb the third doc link is in
			doc4Col = ""		#used to store what columb the fourth doc link is in
			doc5Col = ""		#used to sotre what columb the fith doc lin is in
			wb = openpyxl.load_workbook(fileLoc)				#open up the xlsx file from the convertion
			sheet = wb.worksheets[0]							#the first sheet contains the un analised data
			remove = string.punctuation + string.whitespace		#set the translation filter to remove punctuation and white space
			for col in sheet[1]:								#read each cell in the first row one at a time
				print(str(col.value).lower().translate(None, remove))												#used for testing
				if((str(col.value).lower().translate(None, remove)) == ("name" or "title")):						#look for the title of name of the design example
					print "Name ", col					
					nameCol = com.findColumnLetter(col)																#set nameCol with the letter value of the column. The letter is determined by using the com.findColumnLetter def
				if((str(col.value).lower().translate(None, remove)) == ("documentation1" or "documentation 1")):	#look for the first documentation link
					print "Documentation 1", col
					doc1Col = com.findColumnLetter(col)																#set doc1Col with the letter value of the column. The letter is determined by using the com.findColumnLetter def
				if((str(col.value).lower().translate(None, remove)) == ("documentation2" or "documentation 2")):	#look for the second documentation link
					print "Documentation 2", col
					doc2Col = com.findColumnLetter(col)																#set doc2Col with the letter value of the column. The letter is determined by using the com.findColumnLetter def
				if((str(col.value).lower().translate(None, remove)) == ("documentation3" or "documentation 3")):	#look for the third documentation link
					print "Documentation 3", col
					doc3Col = com.findColumnLetter(col)																#set doc3Col with the letter value of the column. The letter is determined by using the com.findColumnLetter def
				if((str(col.value).lower().translate(None, remove)) == ("documentation4" or "documentation 4")):	#look for the fourth documentation link
					print "Documentation 4", col
					doc4Col = com.findColumnLetter(col)																#set doc4Col with the letter value of the column. The letter is determined by using the com.findColumnLetter def
				if((str(col.value).lower().translate(None, remove)) == ("documentation5" or "documentation 5")):	#look for the fith documentation link
					print "Documentation 5", col
					doc5Col = com.findColumnLetter(col)																#set doc5Col with the letter value of the column. The letter is determined by using the com.findColumnLetter def
			for row in range(2, sheet.max_row + 1):																	#read the xlsx file until there are no more rows to read
				#The line below appends the name and documentation links (all five of them) to a multidimentional list
				com.documentation.append([sheet[nameCol + str(row)].value, sheet[doc1Col + str(row)].value, sheet[doc2Col + str(row)].value, sheet[doc3Col + str(row)].value, sheet[doc4Col + str(row)].value, sheet[doc5Col + str(row)].value])
				#for the line bleow if there is no documentations links. Append the title of the design example to the com.resultList. com.resultList is used to save any example project that have documentation problems
				if([sheet[doc1Col + str(row)].value, sheet[doc2Col + str(row)].value, sheet[doc3Col + str(row)].value, sheet[doc4Col + str(row)].value, sheet[doc5Col + str(row)].value] == ["-", "-", "-", "-", "-"]):
					com.resultList.append([sheet[nameCol + str(row)].value, "-", "No Documentation"])
			com.condenceUrls()	#run the com.condenceUrls function. This function removes duplicates from the com.resultList and com.documentation (comment contines on the next line)
								#lists. This ensures that each link is only checked once and that each design example is only listed once if it has a link problem.
			
		'''
		Def Name:	requesLink
		*
		Arguments:		com 	com contains the varibles with in the class linkStuff
		*				title	this is the string that contains the title of the referance design
		*				link	sting that contains the url of the documentation to be checked
		*
		Returns:		NA 		all changes are done under the class
		*
		Description:	This def uses the session created by the pypac library to use the proxy to check documentation links
		*				any link that does not return a 201, 202, or 203 is appended to the com.resultList. Additionally if
		*				the def fails to make a connection to the server it will return a connection error and append the link
		*				to the list for further investigation.
		'''
		def requesLink(com, title, link):
			try:
				connection = com.session.get(link)																					#use the pypac session to request a responce from the url. (pypac connects us to the intel proxy)
				if(str(connection) != ("<Response [200]>" or "<Response [201]>" or "<Response [202]>" or "<Response [203]>")):		
					com.resultList.append([title, link, connection])																#if the connection does not return successfuly append the project name, link, and error code to the com.resultList
			except:
				com.resultList.append([title, link, "Connection Error"])															#if the session cannont connect at all. append to the project name, link, and "connection error" to the com.resultList
				print "results: ", len(com.resultList)
				print "connection error"
		
		'''
		Def Name:	writeResults
		*
		Arguments:		com 	com contains the varibles with in the class linkStuff
		*
		Returns:		NA 		all changes are done under the class
		*
		Description:	this def use the class to write the com.resultList to the xlsx output sheet. First it opens the file, trys 
		*				to open the sheet titled DocumentationErrors. if the sheet does not alread exsist it creates it. The def
		*				adds the titles of each colum to each row. A loop then appends everything in the com.resultList to the 
		*				sheet. The def then closes by saving the xlsx file.
		'''
		def writeResults(com):
			printable = string.ascii_letters + string.digits			#not really used anymore... probably will delete soon
			wb = openpyxl.load_workbook(fileLoc)						#open the xlsx file from the convertion
			try:	
				sheet = wb.get_sheet_by_name('DocumentationErrors')		#open the sheet in the xlsx named "DocumentationErrors"
			except KeyError:
				wb.create_sheet(title = 'DocumentationErrors')			#if the sheet "DocumentationErrors" does not exsist create it
				sheet = wb.get_sheet_by_name('DocumentationErrors')
			sheet['A1'] = 'Title'										#set the titles of the first row with Title, Link, and Responce
			sheet['B1'] = 'Link'
			sheet['C1'] = 'Responce'
			for i in range(len(com.resultList)):						#loop though the full com.resultList one row at a time. a row in the list contains the name, link and error code of each failed test
				try:
					sheet['A'+str(i+2)] = com.removeNonASCII(str(com.resultList[i][0]))		#write each of the names, links, and error codes to the xlsx file
					sheet['B'+str(i+2)] = com.removeNonASCII(str(com.resultList[i][1]))
					sheet['C'+str(i+2)] = com.removeNonASCII(str(com.resultList[i][2]))
				except:
					print "ASCII error", i													#if the line fails to print due to special charicters return an error and continue writing the file
			wb.save(fileLoc) #add a try here later for error checking
	
	print "creating class"		#used for testing probably will delete soon
	checking = linkStuff()		#set up and run the class linkStuff

'''***********************************************************************************************************************'''
'''********************************************* GUI CLASS LOGIC *********************************************************'''
'''***********************************************************************************************************************'''

class Application(Frame):	
	def __init__(self, master):
		"""Inits the frame"""
		Frame.__init__(self, master)
		self.grid()
		self.inputFileLoc = "None Selected"
		self.saveLoc = "None Selected"
		self.alerts = "None Selected"
		self.create_widgets()
		self.updateInputText()
	
	'''
	Def Name:		create_widgets
	*
	Arguments:		self 	self contains the varibles with in the class Application
	*
	Returns:		NA 		all changes are done under the class
	*
	Description:	This def is used to initial and create the main GUI. Each item in the GUI has its own decloration and name
	'''
	def create_widgets(self):
		#inputInstructs
		#creates text that tells the user the firts instruction
		self.inputInstruct = Label(self, text = "Please browse to the CSV downloaded from the design store.")
		self.inputInstruct.grid(row = 0, column = 0, columnspan = 2, sticky = W)
		
		#Input Browse Button
		#creates a button that opens the input file dialog
		self.inputButton = Button(self, text = "...")
		self.inputButton["command"] = self.openInputFile
		self.inputButton.grid(row = 1, column = 6, sticky = W)
		
		#Input File Text Field
		#creates a text field that shows the input file
		self.inputFileText = Text(self, width = 50, height = 1, wrap = WORD)
		self.inputFileText.grid(row = 1, column = 0, columnspan = 5)
		
		#Save location instructions
		#creates text that tells the user the second instruction
		self.saveInstruct = Label(self, text = "Please browse the location for the processed file to be saved")
		self.saveInstruct.grid(row = 2, column = 0, columnspan = 2, sticky = W)
		
		#Save Location Button
		#creates a button that opens output file dialog
		self.saveButton = Button(self, text = "...")
		self.saveButton["command"] = self.openSaveFile
		self.saveButton.grid(row = 3, column = 6, sticky = W)
		
		#Save Location Text Field
		#creates a text field that displays the output file location
		self.saveLocText = Text(self, width = 50, height = 1, wrap = WORD)
		self.saveLocText.grid(row = 3, column = 0, columnspan = 5)
		
		#Option analytics
		#Creates a checkbox that enables the analytics portion of the script to run
		self.analytics = BooleanVar()
		Checkbutton(self, text = "Referance Design Analytics", variable = self.analytics, command = self.updateInputText).grid(row = 4, column = 0, sticky = W)
		
		#Option check two
		#Creates a checkbox that enables the Documentation link checker portion of the script to run
		self.checkDocs = BooleanVar()
		Checkbutton(self, text = "Check Documentation", variable = self.checkDocs, command = self.updateInputText).grid(row = 5, column = 0, sticky = W)
		
		#run Button
		#creates a button that launches the analytics and documentation if those options are checked
		self.saveButton = Button(self, text = "RUN FOREST RUN!")
		self.saveButton["command"] = self.run
		self.saveButton.grid(row = 6, column = 0, sticky = W)
		
		#Alerts field
		#creates a text field for information to be displayed to the user
		self.alertField = Text(self, width = 50, height = 40, wrap = WORD)
		self.alertField.grid(row = 7, column = 0, columnspan = 5)
		
	'''
	Def Name:		openInputFile
	*
	Arguments:		self 	self contains the varibles with in the class Application
	*
	Returns:		NA 		all changes are done under the class
	*
	Description:	This def is used to get the location and name of the users input CSV or xlsx. The def uses the built in tinkner function askopenfilename
	*				to get the location and name. This def also uses the updateInputText def to refresh the information in all text fields of the gui.
	'''
	def openInputFile(self):
		#the line below sets the inputFileLoc string using the tinkner askopenfilename function. The default input file type is .csv
		self.inputFileLoc = tkFileDialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("CSV files","*.csv"),("XLSX Files", "*.xlsx"),("all files","*.*")))
		self.updateInputText()	#update all text fields in the gui
	
	'''
	Def Name:		openSaveFile
	*
	Arguments:		self 	self contains the varibles with in the class Application
	*
	Returns:		NA 		all changes are done under the class
	*
	Description:	This def is used to get the location the user wants to save the output file. The def uses the built in tinkner function askdirectory
	*				to get the location. The name of the output file is also generated by adding the the current date to _Processed_File.xlsx This def 
	*				also uses the updateInputText def to refresh the information in all text fields of the gui.
	'''
	def openSaveFile(self):
		#the line below sets the saveLoc string using the tinkner askdirectory function. This also generates the name for the file by adding the date to _Processed_File.xlsx
		self.saveLoc = tkFileDialog.askdirectory() + "/" + str(datetime.date.today()) + "_Processed_File.xlsx"
		self.updateInputText()	#update all text fields in the gui
	
	'''
	Def Name:		updateInputText
	*
	Arguments:		self 	self contains the varibles with in the class Application
	*
	Returns:		NA 		all changes are done under the class
	*
	Description:	This def refreshes the text in all text fields with in the gui. This def should not be used if you want to update the text in the alerts field
	*				This def clears all fields before it updates them.
	'''
	def updateInputText(self):
		self.alerts = ""																#Clear any previous messeges in the alerts string
		self.inputFileText.delete(0.0, END)												#Clear the input file text field
		self.inputFileText.insert(0.0, self.inputFileLoc)								#Update the input file text field with the string in inputFileLoc
		self.saveLocText.delete(0.0, END)												#clear the save location text field
		self.saveLocText.insert(0.0, self.saveLoc)										#Update the save location text field with the string saveLoc
		if(self.analytics.get()):
			self.alerts = self.alerts + "INFO: Analytics will run.\n"					#If the analytics check box is checked add the info to the alerts string
		if(self.checkDocs.get()):
			self.alerts = self.alerts + "INFO: Documentation URLs will be checked.\n"	#If the check links check box is checked add the info to the alerts string
		if(self.alerts == ""):
			self.alerts = "*** ERROR: No Option Selected ***"							#If none of the boxes are checked 
		self.alertField.delete(0.0, END)												#Clear the alerts text field of any text
		self.alertField.insert(0.0, self.alerts)										#Display the string alerts in the alerts text field
	
	'''
	Def Name:		addText
	*
	Arguments:		self 	self contains the varibles with in the class Application
	*				message	the string that needs to be added to the alert text field
	*
	Returns:		NA 		all changes are done under the class
	*
	Description:	This def refreshes the text in the alerts text field in addition to adding any message. This should be used for adding text like INFO:
	*				and Errors to the alerts text field
	'''
	def addText(self, message):
		self.alerts = self.alerts + message + "\n"	#add the message sting to the alerts sting
		self.alertField.delete(0.0, END)			#clear the alerts text field
		self.alertField.insert(0.0, self.alerts)	#display the alerts sting in the alerts text field
	
	'''
	Def Name:		run
	*
	Arguments:		self 	self contains the varibles with in the class Application
	*
	Returns:		NA 		all changes are done under the class
	*
	Description:	This def is started by clicking the run button. This def launches the analytics script and the linkCheck script baised off of the checkboxes
	*				with in the gui. There are many error codes that will display if there are problems running the other function the def calls. This def also
	*				includes a class with in it to start a scond GUI that operates the abort button. This class and gui is only launched if the check links option
	*				is selected
	'''
	def run(self):
		if(self.analytics.get()):															#if the analytics check box is checked run the convertion then the analytics
			if (("xlsx" or "XLSX") in self.inputFileLoc):									#check if the input file is formated as an xlsx
				print "already in xlsx"														
				self.addText("INFO: file already in xlsx format")
				if(copyXLSX(self.inputFileLoc, self.saveLoc) == False):						#copy and rename the file to the ouput location. If it fails let the user know.
					self.addText("*** ERROR: Can Not Write to Output Directory ***\nINFO: Try checking if the output directory is writeable")
				else:
					self.addText("INFO: Copy Done")
			elif(("csv" or "CSV") in self.inputFileLoc):									#check if the input file is in CSV format
				print "CSV"
				self.addText("INFO: File in csv format\nINFO: Starting convertion")
				if(convertToXlsx(self.inputFileLoc, self.saveLoc) == False):				#convert and rename the csv file to xlsx format. If it fails alert the user
					self.addText("*** ERROR: Can Not Convert CSV to XLSX! ***\nINFO: Try checking if the output directory is writeable")
				else:
					self.addText("INFO: Convertion done")
			else:
				self.addText("*** ERROR: Invalid File Format ***")
				print "Invalid File Format"
			self.addText("INFO: Starting Analytics")
			runAnalytics(self.saveLoc)														#After the convertion is done run the analytics script.
			self.addText("INFO: Finished Running Analytics")
		if(self.checkDocs.get()):															#If the check documentation box is checked. launch the abort button gui then run the check links def
			self.addText("INFO: Starting link check")
			'''
			subApp class is used to launch the seperate but nested abort button gui.
			'''
			class subApp(Frame):	
				def __init__(sub, master):
					"""Inits the frame"""
					Frame.__init__(sub, master)
					sub.timeCount = 0
					sub.grid()
					sub.create()
				
				def create(sub):
					sub.abortButton = Button(sub, text = "ABORT")
					sub.abortButton["command"] = sub.abortFunction
					sub.abortButton.grid(row = 0, column = 0, sticky = W)
					
				def abortFunction(sub):
					print "set bool"
					bool.set()
					abort.destroy()
			
			print "runningLinks"
			bool = threading.Event()															#creates a flag named bool that can comunicate between threds
			abort = Tk()																		#creates a new gui called abort for the abort button/window
			abort.title("Abort")																#Title the new gui window abort
			abort.geometry("200x50")															#set the size of the abort window
			app2 = subApp(abort)																#get ready to launch the gui
			t = threading.Thread(target = abort.mainloop)										#add the second gui to the thread count
			t = threading.Thread(target = linkCheck, args = (self.saveLoc, bool))				#add the check links to a seperate thread
			t.start()																			#launch the abort gui and the linkCheck script in a seperate thread
		#print "Run Forest Run!\n"

'''***********************************************************************************************************************'''
'''********************************************* MAIN LOGIC CALL *********************************************************'''
'''***********************************************************************************************************************'''
		
gui = Tk()									#new gui named gui
gui.title("Design Store Analytics")			#title the window "Design Store Analytics"
gui.geometry("600x450")						#set the size of the window

app = Application(gui)						#get ready to launch the gui

gui.mainloop()								#launch the gui